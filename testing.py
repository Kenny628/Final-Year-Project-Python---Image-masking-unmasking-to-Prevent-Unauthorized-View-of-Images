import cv2
import numpy as np
import sys
import time
import math
from numpy import fft
import matplotlib.pyplot as plt
from Crypto.Cipher import AES, Blowfish
from Crypto.Util.Padding import pad, unpad
from Crypto.Random import get_random_bytes
from Crypto.Protocol.KDF import PBKDF2
from flask import Flask, request, render_template, send_file
import multiprocessing
from PIL import Image
import pandas as pd
from math import log10, sqrt
import os
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
from skimage.metrics import structural_similarity
import base64
import hashlib
from Crypto import Random
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from stegano import lsb
import numpy as np
from numpy import fft
import math
import cv2
from PIL import Image 
from skimage.metrics import peak_signal_noise_ratio as compare_psnr


def SSIM(original,compressed):
    # before = cv2.imread(original)
    # after = cv2.imread(compressed)
    # Compute SSIM between two images
    (score, diff) = structural_similarity(original, compressed, full=True,channel_axis=2)
    return score
# pip install openpyxl
ivSize = AES.block_size
app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
string_data = "Hello, World!"
encoding_type = "utf-8"  # You can choose the encoding type you want

# Encoding the string to bytes
bytes_data = string_data.encode(encoding_type)
# 仿真运动模糊

def motion_process(image_size, motion_angle):
    PSF = np.zeros(image_size)
    print(image_size)
    center_position = (image_size[0] - 1) / 2
    print(center_position)

    slope_tan = math.tan(motion_angle * math.pi / 180)
    slope_cot = 1 / slope_tan
    if slope_tan <= 1:
        for i in range(15):
            offset = round(i * slope_tan)  # ((center_position-i)*slope_tan)
            PSF[int(center_position + offset), int(center_position - offset)] = 1
        return PSF / PSF.sum()             # 对点扩散函数进行归一化亮度
    else:
        for i in range(15):
            offset = round(i * slope_cot)
            PSF[int(center_position - offset), int(center_position + offset)] = 1
        return PSF / PSF.sum()

# 对图片进行运动模糊
def make_blurred(input, PSF):
    input_fft = fft.fft2(input)            
    PSF_fft = fft.fft2(PSF) 
    blurred = fft.ifft2(input_fft * PSF_fft)
    blurred = np.abs(fft.fftshift(blurred))
    return blurred


def inverse(input, PSF):                
    input_fft = fft.fft2(input)
    PSF_fft = fft.fft2(PSF)          
    result = fft.ifft2(input_fft / PSF_fft)  
    result = np.abs(fft.fftshift(result))
    return result


def normal(array):
    array = np.where(array < 0,  0, array)
    array = np.where(array > 255, 255, array)
    array = array.astype(np.int16)
    return array

def is_32_bit_png(image_path):
    try:
        image = Image.open(image_path)
        if image.mode == "RGBA" and image.getextrema()[3] != (0, 255):
            return True
        return False
    except Exception as e:
        print("Error:", e)
        return False

def convert_32bit_to_24bit(input_path, output_path):
    # Open the 32-bit PNG image
    img = Image.open(input_path)

    # Ensure that the image has an alpha channel
    img = img.convert("RGBA")

    # Create a new image with 24-bit RGB format
    new_img = Image.new("RGB", img.size)

    # Paste the RGB content from the original image to the new image
    new_img.paste(img, (0, 0), img)

    # Save the new image as a 24-bit PNG
    new_img.save(output_path, format="PNG")


def keyCreation(key):
    block_size = 32
    # Getting a hash value of 256 bit (32 byte)
    key = hashlib.sha256(key.encode()).digest()
    return key

# Initialize Flask app and routes here
# ...

# Initialize a DataFrame to Store Results
results_df = pd.DataFrame(columns=['Image Name','Iteration', 'SSIM', 'PSNR', 'PSNR_DATA_HIDING','Encryption Time', 'Decryption Time'])

num_iterations = 10  # Number of iterations
image_path = 'lena.jpg'  # Provide the path to your test image
password = 'your_password'  # Provide the password used for encryption/decryption
count=0
hkey = keyCreation(password)
for filename in os.listdir('misc/'):
    count=0
    for iteration in range(1, num_iterations + 1):
        block_size = 32
    # Get a random initialization vector
        iv = Random.new().read(AES.block_size)
        # using Cipher Block Chaining (CBC) Mode
        # hkey = keyCreation(password)
        encryption_suite = AES.new(hkey, AES.MODE_CBC, iv)

        # If it is string convert to byte string before use it
        if isinstance(password, str):
            password = password.encode()

        # Encrypt the random initialize vector added with the padded data
        cipher_data = encryption_suite.encrypt(iv + pad(password, block_size))

        # Convert the cipher byte string to a base64 string to avoid decode padding error
        cipher_data = base64.b64encode(cipher_data).decode()
        print("Cipher text is :", cipher_data)
        # Encrypt the image
        name_without_extension = os.path.splitext(filename)[0]
        start_encryption_time = time.time()
        print(num_iterations)
        if(count==0):
            OriginalImage=cv2.imread('misc_png/'+name_without_extension+".png")
        else:
            OriginalImage=cv2.imread('misc_recovered/'+name_without_extension+".png")
        b_gray, g_gray, r_gray = cv2.split(OriginalImage.copy())
        # imageOrig = cv2.imdecode(np.frombuffer(imageUserInput, np.uint8), cv2.IMREAD_COLOR)
        # result = is_32_bit_png(imageOrig)
        # if result:
        #     print('hi1')
        # else:
        #     print('hi2')
        img_h, img_w = b_gray.shape[:2]
        PSF = motion_process((img_h,img_w), 6000000000000000000000000000000000000000000000000)
        # blurred = np.abs(make_blurred(image, PSF, 1e-3))
        blurred_b = np.abs(make_blurred(b_gray, PSF))
        print(blurred_b)
        blurred_g = np.abs(make_blurred(g_gray, PSF))
        print(blurred_g)
        blurred_r = np.abs(make_blurred(r_gray, PSF))
        print(blurred_r)
        cv2.imwrite("blurred_b_Image.png", blurred_b)
        cv2.imwrite("blurred_g_Image.png", blurred_g)
        cv2.imwrite("blurred_r_Image.png", blurred_r)
        cv2.waitKey(0)
        blurred = cv2.merge([blurred_b, blurred_g, blurred_r])
        cv2.imwrite("blurred_test.png",blurred)
        secret = lsb.hide("blurred_test.png", cipher_data)
        secret.save("blurred.png")
        end_encryption_time = time.time()
        encryption_time = end_encryption_time - start_encryption_time

        # Decrypt the image
        start_decryption_time = time.time()
        # ... Your decryption logic here ...
        block_size = 32
        cipher_data = lsb.reveal("blurred.png")

        # if not cipher_data:
        #     return None

        cipher_data = base64.b64decode(cipher_data)
        # Retrieve the dynamic initialization vector saved
        iv = cipher_data[:AES.block_size]
        # Retrieved the cipher data
        cipher_data = cipher_data[AES.block_size:]
        decryption_suite = AES.new(hkey, AES.MODE_CBC, iv)
        decrypted_data = unpad(
            decryption_suite.decrypt(cipher_data),
            block_size
        )
        
        image2= cv2.imread("blurred.png")
        b_gray2, g_gray2, r_gray2 = cv2.split(image2.copy())
        # PSF_1 =  extension_PSF(b_gray2, PSF_TEMP_1)
        img_h, img_w = b_gray2.shape[:2] 
        PSF = motion_process((img_h, img_w), 6000000000000000000000000000000000000000000000000)   
        result_blurred_b = inverse(b_gray2,PSF)  # 逆滤波
        normal_deblurred_b=normal(result_blurred_b)
        # PSF_2 =  extension_PSF(g_gray2, PSF_TEMP_2)  
        img_h, img_w = g_gray2.shape[:2] 
        # PSF = motion_process((img_h, img_w), 6000000000000000000000000000000000000000000000000)   
        result_blurred_g = inverse(g_gray2,PSF)  # 逆滤波
        normal_deblurred_g=normal(result_blurred_g)
        # PSF_3 =  extension_PSF(image, PSF_TEMP_3)  
        img_h, img_w = r_gray2.shape[:2] 
        # PSF = motion_process((img_h, img_w), 6000000000000000000000000000000000000000000000000)   
        result_blurred_r = inverse(r_gray2,PSF)  # 逆滤波
        normal_deblurred_r=normal(result_blurred_r)
        recovered = cv2.merge([normal_deblurred_b, normal_deblurred_g, normal_deblurred_r])
        cv2.imwrite("recovere_lsb_aes_password.png",recovered)
        cv2.imwrite("misc_recovered/"+name_without_extension+".png", recovered)
        end_decryption_time = time.time()
        decryption_time = end_decryption_time - start_decryption_time
        # Load the original image for comparison
        original_image = cv2.imread('misc_png/'+name_without_extension+".png")

        # Load the recovered image after decryption
        recovered_image = cv2.imread('misc_recovered/'+name_without_extension+".png")

        # Calculate SSIM and PSNR
        # ssim_value = ssim(original_image, recovered_image)
        psnr_value = compare_psnr(original_image, recovered_image)
        psnr_value_data_hiding=compare_psnr(cv2.imread("blurred_test.png"),cv2.imread("blurred.png"))
        ssim_value=SSIM(original_image,recovered_image)
        # Append the results to the DataFrame
        results_df = pd.concat([results_df, pd.DataFrame({
            'Image Name': [name_without_extension],
            'Iteration': [iteration],
            'SSIM': [ssim_value],
            'PSNR': [psnr_value],
            'PSNR_DATA_HIDING': [psnr_value_data_hiding],
            'Encryption Time': [encryption_time],
            'Decryption Time': [decryption_time]
        })], ignore_index=True)
        count=count+1

# Export Results to Excel
excel_filename = 'encryption_decryption_results2.xlsx'
results_df.to_excel(excel_filename, index=False)
# print("Results exported to", excel_filename)
# results_df = pd.read_excel('encryption_decryption_results.xlsx')

# # Group results by Image Name and calculate mean SSIM and PSNR for each image
# grouped_results = results_df.groupby('Image Name').agg({'SSIM': 'mean', 'PSNR': 'mean'})

# # Create a bar graph
# plt.figure(figsize=(10, 6))
# grouped_results.plot(kind='bar', ax=plt.gca())
# plt.title('Mean SSIM and PSNR for Different Images')
# plt.ylabel('Value')
# plt.xlabel('Image Name')
# plt.xticks(rotation=45, ha='right')
# plt.tight_layout()

# # Save the graph as an image or show it
# plt.savefig('results_graph.png')
# # plt.show()  # Uncomment this line to display the graph

# print("Graph generated and saved as 'results_graph.png'")
import openpyxl
from openpyxl.chart import BarChart, Reference

# Load the existing Excel file or create a new one
excel_filename = 'encryption_decryption_results2.xlsx'
workbook = openpyxl.load_workbook(excel_filename)

# Access the active worksheet (assuming the results are on the first sheet)
worksheet = workbook.active

# Group results by Image Name and calculate mean SSIM and PSNR for each image
grouped_results = results_df.groupby('Image Name').agg({'SSIM': 'mean', 'PSNR': 'mean'})

# Create a new sheet for the chart
chart_sheet = workbook.create_sheet(title='Graph')

# Add data to the chart
data = Reference(worksheet,
                 min_col=worksheet.max_column - 1,  # Column with Image Names
                 min_row=2, max_row=grouped_results.shape[0] + 1)
categories = Reference(worksheet,
                       min_col=worksheet.max_column,  # Column with SSIM and PSNR
                       min_row=1, max_row=1)
chart = BarChart()
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Mean SSIM and PSNR for Different Images"
chart.x_axis.title = "Image Name"
chart.y_axis.title = "Value"
chart.width = 12  # Adjust the chart width

# Add the chart to the chart sheet
chart_sheet.add_chart(chart, "A1")

# Save the updated Excel file
workbook.save(excel_filename)
print("Chart added to the Excel file:", excel_filename)

